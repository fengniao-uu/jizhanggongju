from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple
import csv
import datetime as dt

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

import config
from db import get_adapter
from utils.validators import (
    is_amount_positive,
    is_amount_nonnegative,
    is_yyyymmdd,
    safe_str,
    type_in,
    status_in,
)

TX_HEADERS = ["日期", "类型", "分类", "金额", "房间号", "描述", "标签"]
REM_HEADERS = ["房间号", "房租金额", "到期日期", "租期结束日期", "状态", "备注", "提醒标签"]


def _today_str() -> str:
    return dt.date.today().isoformat()


class IOService:
    """导出 xlsx/csv + 导入解析预览/确认写入"""

    @staticmethod
    def _ok(data=None, msg="ok", code=0):
        return {"code": code, "msg": msg, "data": data}

    @staticmethod
    def _fail(msg, code=400, data=None):
        return {"code": code, "msg": msg, "data": data}, code

    # ========== 导出：交易 ==========
    @staticmethod
    def export_transactions(items: List[Dict], fmt: str = "xlsx") -> Tuple[BytesIO | StringIO, str, str]:
        rows: List[List[Any]] = [list(TX_HEADERS)]
        for it in items:
            rows.append([
                safe_str(it.get("trans_date") or it.get("date")),
                safe_str(it.get("type")),
                safe_str(it.get("category")),
                float(it.get("amount") or 0),
                safe_str(it.get("room_no") or it.get("room")),
                safe_str(it.get("description") or it.get("note")),
                safe_str(it.get("tag") or ""),
            ])
        if fmt == "csv":
            buf = StringIO()
            writer = csv.writer(buf)
            for r in rows:
                writer.writerow([("" if v is None else v) for v in r])
            buf.seek(0)
            fname = f"transactions_{_today_str()}.csv"
            mime = "text/csv; charset=utf-8-sig"
            return buf, fname, mime
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl 未安装，无法导出 Excel 文件")
        buf = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "交易记录"
        widths = [14, 10, 14, 12, 14, 40, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if r_idx == 1:
                    c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        wb.save(buf)
        buf.seek(0)
        fname = f"transactions_{_today_str()}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return buf, fname, mime

    # ========== 导出：提醒 ==========
    @staticmethod
    def export_reminders(items: List[Dict], fmt: str = "xlsx") -> Tuple[BytesIO | StringIO, str, str]:
        rows: List[List[Any]] = [list(REM_HEADERS)]
        for it in items:
            rows.append([
                safe_str(it.get("room_no") or it.get("room")),
                float(it.get("rent_amount") or it.get("amount") or 0),
                safe_str(it.get("due_date") or ""),
                safe_str(it.get("lease_end_date") or it.get("end_date") or ""),
                safe_str(it.get("status") or "未完成"),
                safe_str(it.get("note") or it.get("remark") or ""),
                safe_str(it.get("smart_tag") or ""),
            ])
        if fmt == "csv":
            buf = StringIO()
            writer = csv.writer(buf)
            for r in rows:
                writer.writerow([("" if v is None else v) for v in r])
            buf.seek(0)
            fname = f"reminders_{_today_str()}.csv"
            mime = "text/csv; charset=utf-8-sig"
            return buf, fname, mime
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl 未安装，无法导出 Excel 文件")
        buf = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "收租提醒"
        widths = [14, 12, 14, 16, 10, 40, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if r_idx == 1:
                    c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        wb.save(buf)
        buf.seek(0)
        fname = f"reminders_{_today_str()}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return buf, fname, mime

    # ========== 通用功能：读取上传的 Excel/CSV 二维数据 ==========
    @staticmethod
    def _read_rows(file_bytes: bytes, filename: str) -> List[List[Any]]:
        fn = (filename or "").lower()
        if fn.endswith(".csv"):
            text = file_bytes.decode("utf-8-sig", errors="ignore")
            reader = csv.reader(StringIO(text))
            return [list(x) for x in reader]
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl 未安装，无法读取 Excel 文件")
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row or []))
        wb.close()
        return rows

    # ========== 导入：交易预览 + 确认 ==========
    @staticmethod
    def preview_transactions(user: Dict, file_bytes: bytes, filename: str) -> Dict:
        user_id = int(user["id"])
        try:
            rows = IOService._read_rows(file_bytes, filename)
        except Exception as e:
            return IOService._fail(f"无法解析文件：{e}", 400)
        if not rows:
            return IOService._fail("文件为空", 400)
        header = [str(x).strip() for x in rows[0]]
        body = rows[1:]

        def _idx(names: List[str]) -> int:
            for n in names:
                for i, h in enumerate(header):
                    if h and n in h:
                        return i
            return -1

        i_date = _idx(["日期", "date", "trans_date"])
        i_type = _idx(["类型", "type"])
        i_cat = _idx(["分类", "category"])
        i_amt = _idx(["金额", "amount", "money"])
        i_room = _idx(["房间号", "room_no", "room", "房间"])
        i_desc = _idx(["描述", "description", "note", "备注"])
        i_tag = _idx(["标签", "tag"])

        cats_by_type = get_adapter().list_categories_grouped(user_id)
        income_cats = {c["name"] for c in cats_by_type.get("收入", [])}
        expense_cats = {c["name"] for c in cats_by_type.get("支出", [])}

        ok_items: List[Dict] = []
        errors: List[Dict] = []
        for r_idx, raw in enumerate(body, start=2):
            line = {"行号": r_idx}
            if not any((v is not None and str(v).strip() != "") for v in raw):
                continue
            s_type = safe_str(raw[i_type]) if i_type >= 0 else ""
            for alias_zh, alias_en, canonical in (
                ("收入", "income", "收入"),
                ("支出", "expense", "支出"),
            ):
                if alias_zh in s_type or alias_en in s_type.lower():
                    s_type = canonical
                    break
            if not type_in(s_type):
                errors.append({**line, "错误": f"类型无效（{s_type or '空'}），请填「收入/支出」"})
                continue
            amt_raw = raw[i_amt] if i_amt >= 0 else None
            if not is_amount_positive(amt_raw):
                errors.append({**line, "错误": "金额必须 > 0"})
                continue
            cat = safe_str(raw[i_cat]) if i_cat >= 0 else ""
            allowed = income_cats if s_type == "收入" else expense_cats
            if cat not in allowed:
                errors.append({**line, "错误": f"分类「{cat}」不属于 {s_type}（可选 {sorted(allowed)[:10]}）"})
                continue
            td = safe_str(raw[i_date]) if i_date >= 0 else ""
            if td and not is_yyyymmdd(td):
                errors.append({**line, "错误": f"日期格式错误（{td}），应为 YYYY-MM-DD"})
                continue
            if not td:
                td = _today_str()
            row = {
                "type": s_type,
                "category": cat,
                "amount": float(amt_raw),
                "room_no": (safe_str(raw[i_room]) if i_room >= 0 else "")[:32],
                "description": (safe_str(raw[i_desc]) if i_desc >= 0 else "")[:255],
                "tag": (safe_str(raw[i_tag]) if i_tag >= 0 else "")[:64],
                "trans_date": td,
            }
            ok_items.append({**row, "_line": r_idx})
        return IOService._ok(
            {
                "total": len(body),
                "ok": len(ok_items),
                "errors": errors,
                "rows": ok_items[:200],
                "_ok_rows_full": ok_items,
                "sample_count": len(ok_items[:200]),
            },
            f"预览完成：可导入 {len(ok_items)} 条，错误 {len(errors)} 条",
        )

    @staticmethod
    def confirm_transactions(user: Dict, preview_rows: List[Dict]) -> Dict:
        user_id = int(user["id"])
        if not isinstance(preview_rows, list) or not preview_rows:
            return IOService._fail("没有可导入的预览数据", 400)
        adapter = get_adapter()
        cats_by_type = adapter.list_categories_grouped(user_id)
        income_cats = {c["name"] for c in cats_by_type.get("收入", [])}
        expense_cats = {c["name"] for c in cats_by_type.get("支出", [])}
        inserted = 0
        skipped = 0
        ids: List[int] = []
        for raw in preview_rows:
            if not isinstance(raw, dict):
                skipped += 1
                continue
            row = {}
            t = safe_str(raw.get("type"))
            if not type_in(t):
                skipped += 1
                continue
            allowed = income_cats if t == "收入" else expense_cats
            cat = safe_str(raw.get("category"), 64)
            if cat not in allowed:
                skipped += 1
                continue
            if not is_amount_positive(raw.get("amount")):
                skipped += 1
                continue
            td = safe_str(raw.get("trans_date"))
            if td and not is_yyyymmdd(td):
                skipped += 1
                continue
            if not td:
                td = _today_str()
            row["type"] = t
            row["category"] = cat
            row["amount"] = float(raw["amount"])
            row["trans_date"] = td
            row["room_no"] = safe_str(raw.get("room_no"), 32)
            row["description"] = safe_str(raw.get("description"), 255)
            row["tag"] = safe_str(raw.get("tag"), 64)
            try:
                tx_id = adapter.create_transaction(user_id, row)
                ids.append(int(tx_id))
                inserted += 1
            except Exception:
                skipped += 1
        return IOService._ok(
            {"inserted": inserted, "skipped": skipped, "ids": ids[:500]},
            f"导入完成：成功 {inserted} 条，跳过 {skipped} 条",
        )

    # ========== 导入：提醒预览 + 确认 ==========
    @staticmethod
    def preview_reminders(user: Dict, file_bytes: bytes, filename: str) -> Dict:
        try:
            rows = IOService._read_rows(file_bytes, filename)
        except Exception as e:
            return IOService._fail(f"无法解析文件：{e}", 400)
        if not rows:
            return IOService._fail("文件为空", 400)
        header = [str(x).strip() for x in rows[0]]
        body = rows[1:]

        def _idx(names: List[str]) -> int:
            for n in names:
                for i, h in enumerate(header):
                    if h and n in h:
                        return i
            return -1

        i_room = _idx(["房间号", "room_no", "room", "房间"])
        i_amt = _idx(["房租金额", "rent_amount", "amount", "金额"])
        i_due = _idx(["到期日期", "due_date"])
        i_end = _idx(["租期结束日期", "end_date", "lease_end", "lease_end_date"])
        i_st = _idx(["状态", "status"])
        i_note = _idx(["备注", "note", "remark"])

        ok_items: List[Dict] = []
        errors: List[Dict] = []
        for r_idx, raw in enumerate(body, start=2):
            line = {"行号": r_idx}
            if not any((v is not None and str(v).strip() != "") for v in raw):
                continue
            room = (safe_str(raw[i_room]) if i_room >= 0 else "")[:32]
            if not room:
                errors.append({**line, "错误": "房间号不能为空"})
                continue
            amt_raw = raw[i_amt] if i_amt >= 0 else None
            if not is_amount_nonnegative(amt_raw):
                errors.append({**line, "错误": "房租金额必须 >= 0"})
                continue
            due = safe_str(raw[i_due]) if i_due >= 0 else ""
            end = safe_str(raw[i_end]) if i_end >= 0 else ""
            if not is_yyyymmdd(due):
                errors.append({**line, "错误": f"到期日期格式错误（{due}），应为 YYYY-MM-DD"})
                continue
            if end and not is_yyyymmdd(end):
                errors.append({**line, "错误": f"租期结束日期格式错误（{end}）"})
                continue
            st = safe_str(raw[i_st]) if i_st >= 0 else "未完成"
            if st and not status_in(st):
                for alias, canon in (("未完成", "未完成"), ("待确认", "未完成"), ("已交", "已完成"), ("已完成", "已完成"), ("已确认", "已确认")):
                    if alias in st:
                        st = canon
                        break
                if not status_in(st):
                    st = "未完成"
            row = {
                "room_no": room,
                "rent_amount": float(amt_raw or 0),
                "due_date": due,
                "lease_end_date": end or "",
                "status": st or "未完成",
                "note": (safe_str(raw[i_note]) if i_note >= 0 else "")[:255],
            }
            ok_items.append({**row, "_line": r_idx})
        return IOService._ok(
            {
                "total": len(body),
                "ok": len(ok_items),
                "errors": errors,
                "rows": ok_items[:200],
                "_ok_rows_full": ok_items,
                "sample_count": len(ok_items[:200]),
            },
            f"预览完成：可导入 {len(ok_items)} 条，错误 {len(errors)} 条",
        )

    @staticmethod
    def confirm_reminders(user: Dict, preview_rows: List[Dict]) -> Dict:
        user_id = int(user["id"])
        if not isinstance(preview_rows, list) or not preview_rows:
            return IOService._fail("没有可导入的预览数据", 400)
        adapter = get_adapter()
        inserted = 0
        skipped = 0
        ids: List[int] = []
        for raw in preview_rows:
            if not isinstance(raw, dict):
                skipped += 1
                continue
            room = safe_str(raw.get("room_no"), 32)
            amt_raw = raw.get("rent_amount")
            if not room or not is_amount_nonnegative(amt_raw):
                skipped += 1
                continue
            due = safe_str(raw.get("due_date"))
            end = safe_str(raw.get("lease_end_date"))
            if not is_yyyymmdd(due):
                skipped += 1
                continue
            if end and not is_yyyymmdd(end):
                skipped += 1
                continue
            st = safe_str(raw.get("status")) or "未完成"
            if not status_in(st):
                st = "未完成"
            row = {
                "room_no": room,
                "rent_amount": float(amt_raw),
                "due_date": due,
                "lease_end_date": end,
                "status": st,
                "note": safe_str(raw.get("note"), 255),
            }
            try:
                rid = adapter.create_reminder(user_id, row)
                ids.append(int(rid))
                inserted += 1
            except Exception:
                skipped += 1
        return IOService._ok(
            {"inserted": inserted, "skipped": skipped, "ids": ids[:500]},
            f"导入完成：成功 {inserted} 条，跳过 {skipped} 条",
        )

    # ========== 全量备份：用户 + 交易 + 提醒 + 分类（dict、JSON 序列化友好格式） ==========
    @staticmethod
    def backup_user(user: Dict) -> Dict[str, Any]:
        user_id = int(user["id"])
        adapter = get_adapter()
        cats = adapter.list_categories_grouped(user_id)
        tx_raw = adapter.list_transactions(user_id, {}, 1, 200_000)
        if isinstance(tx_raw, dict):
            tx = tx_raw.get("items") or tx_raw.get("list") or []
        else:
            tx = list(tx_raw or [])
        rem_raw = adapter.list_reminders(user_id, {})
        if isinstance(rem_raw, dict):
            rems = rem_raw.get("items") or rem_raw.get("list") or []
        else:
            rems = list(rem_raw or [])
        return {
            "meta": {
                "version": config.DB_VERSION,
                "exported_at": dt.datetime.now().isoformat(timespec="seconds"),
                "db_adapter": config.DB_ADAPTER,
            },
            "user": {
                "user_id": user_id,
                "account_no": user.get("account_no"),
                "created_at": str(user.get("created_at") or ""),
            },
            "categories": cats,
            "transactions": tx,
            "reminders": rems,
        }
